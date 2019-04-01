from bs4 import BeautifulSoup
import requests
from datetime import datetime
from dateutil import parser
from openpyxl import Workbook

session = requests.session()


def fetch_url(url):
    req = session.get(url)
    if req.status_code != 200:
        raise Exception("Unable to download: {}: status_code: {}".format(url, req.status_code))
    soup = BeautifulSoup(req.text, "html.parser")
    return soup


def get_data(url):

    data = fetch_url(url)
    table = data.find('table')
    trs = table.find_all('tr')
    result = []
    last_state = ''
    for tr in trs[1:]:
        tds = tr.find_all('td')
        state = tds[0].text
        constituencies = tds[1].text.split(',')
        date = tds[2].text.split('(')[0] + " 2019"
        if date[:4] == 'Apri' and date[4]!= 'l' :
            date = 'April 29 2019'
        d = parser.parse(date)
        date = d
        for con in constituencies:
            if len(state) == 0:
                state = last_state
            li = [state, con, date]
            result.append(li)
        last_state = state

    workbook = Workbook()
    ws = workbook.active
    row_number = 3
    for res in result:
        ws.cell(row= row_number, column=1).value = res[0].strip()
        ws.cell(row= row_number, column=2).value = res[1].strip()
        ws.cell(row= row_number, column=3).value = res[2].date().isoformat().strip()
        row_number = row_number + 1
    workbook.save('dates.xlsx')






if __name__ == '__main__':
    url = 'https://www.businessinsider.in/lok-sabha-elections-2019-polling-dates-and-schedule-by-state-and-constituency/articleshow/68462841.cms'
    get_data(url)